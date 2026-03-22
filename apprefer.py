import os
import json
import io
import uuid
import time
import shutil
import zipfile
import csv
import traceback
from pathlib import Path
from typing import Optional, List, Dict, Any
from datetime import datetime
from spectrogram_analyzer import extract_enhanced_call_parameters
from fastapi import FastAPI, HTTPException, UploadFile, File, Form, BackgroundTasks
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import librosa
import librosa.display
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
from scipy import signal
from scipy.signal import butter, sosfilt
from supabase import create_client, Client
from groq import Groq
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.colors import black, HexColor
from reportlab.lib.enums import TA_CENTER

from dotenv import load_dotenv
import re

# Import prediction module
from predict import MultiSpeciesPredictor
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# Ensure this path exists in your directories setup
BATCH_DIR = Path("batches")
BATCH_DIR.mkdir(parents=True, exist_ok=True)


def extract_datetime_from_filename(filename: str) -> tuple[Optional[str], Optional[str]]:
    """
    Extract date and time from various filename formats.
    
    Supported formats:
    - 20250318_181801 -> YYYYMMDD_HHMMSS
    - 20211126_182140_5_20211126182140 -> ends with YYYYMMDDHHMMSS
    - BCIT___20110102_230418_20110102230420 -> ends with YYYYMMDDHHMMSS
    
    Returns:
        tuple: (date_str, time_str) in format ("YYYY-MM-DD", "HH:MM:SS") or (None, None)
    """
    # Remove file extension
    name_only = Path(filename).stem
    
    # Pattern 1: YYYYMMDD_HHMMSS (simple format at start)
    # Example: 20250318_181801
    match = re.match(r'^(\d{4})(\d{2})(\d{2})_(\d{2})(\d{2})(\d{2})', name_only)
    if match:
        year, month, day, hour, minute, second = match.groups()
        date_str = f"{year}-{month}-{day}"
        time_str = f"{hour}:{minute}:{second}"
        return date_str, time_str
    
    # Pattern 2 & 3: Ends with YYYYMMDDHHMMSS (14 digits)
    # Example: 20211126_182140_5_20211126182140 or BCIT___20110102_230418_20110102230420
    match = re.search(r'(\d{14})$', name_only)
    if match:
        timestamp = match.group(1)
        year = timestamp[0:4]
        month = timestamp[4:6]
        day = timestamp[6:8]
        hour = timestamp[8:10]
        minute = timestamp[10:12]
        second = timestamp[12:14]
        date_str = f"{year}-{month}-{day}"
        time_str = f"{hour}:{minute}:{second}"
        return date_str, time_str
    
    # Pattern 4: Alternative format with underscore (YYYYMMDD_HHMMSS somewhere in filename)
    match = re.search(r'(\d{8})_(\d{6})', name_only)
    if match:
        date_part, time_part = match.groups()
        year = date_part[0:4]
        month = date_part[4:6]
        day = date_part[6:8]
        hour = time_part[0:2]
        minute = time_part[2:4]
        second = time_part[4:6]
        date_str = f"{year}-{month}-{day}"
        time_str = f"{hour}:{minute}:{second}"
        return date_str, time_str
    
    # No match found
    return None, None


# Update Batch Result to store metadata

# ============================================
# CONFIGURATION
# ============================================
app = FastAPI(title="Bat Call Analyzer API", version="3.0")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ──────────────────────────────────
# Directories
# ──────────────────────────────────
load_dotenv()

UPLOADS_DIR      = Path("uploads")
SPECTROGRAMS_DIR = Path("spectrograms")
TEMP_DIR         = Path("temp")
RESULTS_DIR      = Path("results")
STATIC_DIR       = Path("static/bat_species")

for d in [UPLOADS_DIR, SPECTROGRAMS_DIR, TEMP_DIR, RESULTS_DIR, STATIC_DIR]:
    d.mkdir(parents=True, exist_ok=True)

# ──────────────────────────────────
# Supabase initialisation
# ──────────────────────────────────
SUPABASE_URL        = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY")
SUPABASE_BUCKET     = os.getenv("SUPABASE_BUCKET", "bcit-bat")

supabase: Optional[Client] = None
SUPABASE_ENABLED = False

try:
    if not SUPABASE_URL or not SUPABASE_SERVICE_KEY:
        print("⚠ SUPABASE_URL or SUPABASE_SERVICE_KEY missing – cloud storage disabled")
    else:
        supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)
        SUPABASE_ENABLED = True
        print(f"✓ Supabase client connected")
        print(f"✓ Storage bucket: {SUPABASE_BUCKET}")
        print(f"✓ Database table: analysis_results")
except Exception as e:
    print(f"⚠ Supabase initialisation failed: {e}")
    SUPABASE_ENABLED = False

# ──────────────────────────────────
# Groq
# ──────────────────────────────────
GROQ_API_KEY = os.getenv("GROQ_API_KEY")
groq_client  = Groq(api_key=GROQ_API_KEY) if GROQ_API_KEY else None

# ──────────────────────────────────
# ML Model
# ──────────────────────────────────
MODEL_PATH   = os.getenv("MODEL_PATH", "models/efficientnet_b0_bat_3_dataset(1).pth")
CLASSES_PATH = os.getenv("CLASSES_PATH", "models/new_3_dataset_classes(1).json")
predictor    = MultiSpeciesPredictor(MODEL_PATH, CLASSES_PATH)

# ──────────────────────────────────
# Spectrogram Themes
# ──────────────────────────────────
SPECTROGRAM_THEMES = {
    "dark_viridis":      {"cmap": "viridis", "bg_color": "black", "tick_color": "white"},
    "bright_plasma":     {"cmap": "plasma",  "bg_color": "white", "tick_color": "black"},
    "classic_grayscale": {"cmap": "gray",    "bg_color": "white", "tick_color": "black"},
    "inferno":           {"cmap": "inferno", "bg_color": "black", "tick_color": "white"},
    "magma":             {"cmap": "magma",   "bg_color": "black", "tick_color": "white"},
    "jet":               {"cmap": "jet",     "bg_color": "white", "tick_color": "black"},
}

# ============================================
# PYDANTIC MODELS
# ============================================
class SpeciesDetection(BaseModel):
    species:    str
    confidence: float
    rank:       int

class CallParameters(BaseModel):
    start_frequency: float
    end_frequency:   float
    peak_frequency:  float
    bandwidth:       float
    intensity:       float
    pulse_duration:  float
    total_length:    float
    shape:           str

class AnalysisResult(BaseModel):
    file_id:            str
    original_filename:  str
    timestamp:          int
    duration:           float
    sample_rate:        int
    species_detected:   List[SpeciesDetection]
    call_parameters:    CallParameters
    spectrogram_url:    str
    audio_url:          str
    supabase_urls:      Optional[Dict[str, Optional[str]]] = None
    processing_mode:    str
    display_theme:      str
    species_image_url:  Optional[str] = None
    sync_status:        Optional[str] = "pending"

class BatchAnalysisResult(BaseModel):
    batch_id:    str
    total_files: int
    completed:   int
    failed:      int
    results:     List[AnalysisResult]
    created_at:  int
    theme:       str
    input_type:  str

class BatchAnalysisResult(BaseModel):
    batch_id:   str
    total_files: int
    completed:  int
    failed:     int
    results:    List[AnalysisResult]

class AIReportRequest(BaseModel):
    file_ids: List[str]
    query:    Optional[str] = None

class AnalysisResult(BaseModel):
    file_id:            str
    original_filename:  str
    timestamp:          int
    duration:           float
    sample_rate:        int
    species_detected:   List[SpeciesDetection]
    call_parameters:    CallParameters
    spectrogram_url:    str
    audio_url:          str
    supabase_urls:      Optional[Dict[str, Optional[str]]] = None
    processing_mode:    str
    display_theme:      str
    species_image_url:  Optional[str] = None
    sync_status:        Optional[str] = "pending"
    batch_id:           Optional[str] = None  # <--- NEW FIELD

# ============================================
# SUPABASE HELPER FUNCTIONS
# ============================================
def get_species_image_path(species_name: str) -> Optional[Path]:
    """Get local path to species image for Excel/PDF embedding."""
    clean_name = species_name.replace(" ", "_").replace("/", "_")
    for ext in (".jpg", ".jpeg", ".png"):
        local_image = STATIC_DIR / f"{clean_name}{ext}"
        if local_image.exists():
            return local_image
    placeholder = STATIC_DIR / "placeholder.jpg"
    if placeholder.exists():
        return placeholder
    return None

def generate_excel_report(results: List[AnalysisResult], batch_id: Optional[str] = None) -> io.BytesIO:
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bat Call Analysis"
    
    # Styling
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers = [
        "File ID", "Batch ID", "Filename", 
        "Recording Date", "Recording Time",
        "Analysis Date", "Analysis Time",
        "Top Species", "Confidence (%)", "Start Freq (kHz)", 
        "Peak Freq (kHz)", "End Freq (kHz)", "Bandwidth (kHz)",
        "Intensity (dB)", "Pulse Duration (ms)", "Total Length (ms)",
        "Shape", "Species Image"
    ]
    
    # Write Headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Write Data
    for row_num, result in enumerate(results, 2):
        sd = result.species_detected[0] if result.species_detected else None
        cp = result.call_parameters
        
        # Analysis timestamp
        analysis_dt = datetime.fromtimestamp(result.timestamp)
        
        # Extract recording date/time from filename
        recording_date, recording_time = extract_datetime_from_filename(result.original_filename)
        
        row_data = [
            result.file_id,
            result.batch_id or "N/A",
            result.original_filename,
            recording_date or "N/A",
            recording_time or "N/A",
            analysis_dt.strftime("%Y-%m-%d"),
            analysis_dt.strftime("%H:%M:%S"),
            sd.species if sd else "Unknown",
            f"{sd.confidence:.1f}" if sd else "0",
            f"{cp.start_frequency:.1f}",
            f"{cp.peak_frequency:.1f}",
            f"{cp.end_frequency:.1f}",
            f"{cp.bandwidth:.1f}",
            f"{cp.intensity:.1f}",
            f"{cp.pulse_duration:.2f}",
            f"{cp.total_length:.2f}",
            cp.shape,
            "" # Placeholder for image
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = center_align
            cell.border = border

        # Embed Image
        if sd:
            img_path = get_species_image_path(sd.species)
            if img_path:
                try:
                    img = XLImage(str(img_path))
                    # Resize image to fit cell roughly
                    img.width = 80
                    img.height = 80
                    # Anchor image to the cell
                    ws.add_image(img, f"{get_column_letter(len(headers))}{row_num}")
                    # Adjust row height to fit image
                    ws.row_dimensions[row_num].height = 65
                except Exception as e:
                    print(f"Error adding image to Excel: {e}")

    # Adjust Column Widths
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 20
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def _public_url(object_path: str) -> str:
    """Build the public URL for an object in the Supabase bucket."""
    return f"{SUPABASE_URL}/storage/v1/object/public/{SUPABASE_BUCKET}/{object_path}"


def upload_to_supabase_storage(local_path: Path, object_path: str) -> Optional[str]:
    """
    Upload a file to Supabase Storage.
    Returns the public URL on success, None on failure.
    """
    if not SUPABASE_ENABLED:
        return None
    try:
        with open(local_path, "rb") as f:
            file_bytes = f.read()

        # Determine content type
        suffix = local_path.suffix.lower()
        content_type_map = {".wav": "audio/wav", ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg"}
        content_type = content_type_map.get(suffix, "application/octet-stream")

        # FIXED: Use proper header format for Supabase
        supabase.storage.from_(SUPABASE_BUCKET).upload(
            path=object_path,
            file=file_bytes,
            file_options={"content-type": content_type, "x-upsert": "true"},
        )

        url = _public_url(object_path)
        print(f"  ✓ Uploaded: {object_path}")
        return url
    except Exception as e:
        print(f"  ✗ Upload failed for {object_path}: {e}")
        traceback.print_exc()
        return None


def delete_from_supabase_storage(object_path: str) -> bool:
    """Delete a single object from Supabase Storage."""
    if not SUPABASE_ENABLED:
        return False
    try:
        supabase.storage.from_(SUPABASE_BUCKET).remove([object_path])
        print(f"✓ Deleted from Supabase Storage: {object_path}")
        return True
    except Exception as e:
        print(f"❌ Supabase storage delete failed for {object_path}: {e}")
        return False


def save_result_to_supabase(result_dict: dict) -> bool:
    """
    Upsert an analysis-result row into Supabase.
    """
    if not SUPABASE_ENABLED:
        return False
    try:
        row = result_dict.copy()
        row["sync_status"] = "synced"
        row["synced_at"]   = datetime.utcnow().isoformat()

        # Serialise nested Pydantic / dict fields → JSON strings
        for key in ("species_detected", "call_parameters", "supabase_urls"):
            if key in row and not isinstance(row[key], str):
                row[key] = json.dumps(row[key]) if row[key] is not None else None

        supabase.table("analysis_results").upsert(row).execute()
        print(f"  ✓ Supabase row upserted: {row['file_id']}")
        return True
    except Exception as e:
        print(f"  ✗ Supabase upsert failed: {e}")
        traceback.print_exc()
        return False


def get_result_from_supabase(file_id: str) -> Optional[Dict]:
    """Fetch a single result row by file_id."""
    if not SUPABASE_ENABLED:
        return None
    try:
        resp = supabase.table("analysis_results").select("*").eq("file_id", file_id).execute()
        if resp.data:
            row = resp.data[0]
            for key in ("species_detected", "call_parameters", "supabase_urls"):
                if key in row and isinstance(row[key], str):
                    row[key] = json.loads(row[key])
            return row
        return None
    except Exception as e:
        print(f"❌ Supabase select failed for {file_id}: {e}")
        return None


def get_all_results_from_supabase() -> List[Dict]:
    """Fetch every result row, ordered newest-first."""
    if not SUPABASE_ENABLED:
        return []
    try:
        resp = supabase.table("analysis_results").select("*").order("timestamp", desc=True).execute()
        rows = resp.data or []
        for row in rows:
            for key in ("species_detected", "call_parameters", "supabase_urls"):
                if key in row and isinstance(row[key], str):
                    row[key] = json.loads(row[key])
        return rows
    except Exception as e:
        print(f"❌ Supabase select-all failed: {e}")
        return []


def delete_result_from_supabase(file_id: str) -> bool:
    """Delete a single row by file_id."""
    if not SUPABASE_ENABLED:
        return False
    try:
        supabase.table("analysis_results").delete().eq("file_id", file_id).execute()
        print(f"✓ Deleted result {file_id} from Supabase")
        return True
    except Exception as e:
        print(f"❌ Supabase delete failed for {file_id}: {e}")
        return False


# ──────────────────────────────────
# Background upload task
# ──────────────────────────────────
def async_supabase_upload_task(
    file_id: str,
    audio_path: Path,
    slow_audio_path: Path,
    spec_display_path: Path,
    result_dict: dict,
):
    """
    Runs via FastAPI BackgroundTasks – fires AFTER the HTTP response is sent.
    Uploads files → updates URLs in the dict → upserts row → persists local JSON.
    """
    if not SUPABASE_ENABLED:
        print(f"⚠ Supabase disabled, skipping upload for {file_id}")
        return

    print(f"\n📤 [Background Sync] Starting for {file_id}...")
    start_time = time.time()
    result_dict["sync_status"] = "syncing"

    supabase_urls: Dict[str, Optional[str]] = {}
    all_ok = True

    # ── audio (original) ──
    if audio_path.exists():
        url = upload_to_supabase_storage(audio_path, f"audio/{file_id}.wav")
        if url:
            supabase_urls["audio"] = url
        else:
            all_ok = False
    else:
        print(f"  ⚠ Audio file not found: {audio_path}")

    # ── audio (slowed) ──
    if slow_audio_path.exists():
        url = upload_to_supabase_storage(slow_audio_path, f"audio/{file_id}_slow.wav")
        if url:
            supabase_urls["audio_slow"] = url
        else:
            all_ok = False
    else:
        print(f"  ⚠ Slow audio file not found: {slow_audio_path}")

    # ── spectrogram ──
    if spec_display_path.exists():
        url = upload_to_supabase_storage(spec_display_path, f"spectrograms/{file_id}.png")
        if url:
            supabase_urls["spectrogram"] = url
        else:
            all_ok = False
    else:
        print(f"  ⚠ Spectrogram file not found: {spec_display_path}")

    result_dict["supabase_urls"] = supabase_urls

    # ── persist to DB ──
    db_ok = save_result_to_supabase(result_dict)

    # ── final status ──
    result_dict["sync_status"] = "synced" if (all_ok and db_ok) else "failed"

    # ── update local JSON so the UI can read sync_status without hitting the DB ──
    try:
        with open(RESULTS_DIR / f"{file_id}.json", "w") as f:
            json.dump(result_dict, f, indent=2)
    except Exception as e:
        print(f"  ✗ Failed to update local JSON: {e}")

    elapsed = time.time() - start_time
    print(f"✅ [Background Sync] Complete for {file_id} ({elapsed:.2f}s)  status={result_dict['sync_status']}")


# ──────────────────────────────────
# Species image helper (local only)
# ──────────────────────────────────
def get_species_image_url(species_name: str) -> Optional[str]:
    clean_name = species_name.replace(" ", "_").replace("/", "_")
    for ext in (".jpg", ".jpeg", ".png"):
        local_image = STATIC_DIR / f"{clean_name}{ext}"
        if local_image.exists():
            return f"/api/static/bat_species/{local_image.name}"
    placeholder = STATIC_DIR / "placeholder.jpg"
    if placeholder.exists():
        return "/api/static/bat_species/placeholder.jpg"
    return None


# ============================================
# ANALYSIS HELPER FUNCTIONS
# ============================================
def analyze_spectrogram_shape(S_db: np.ndarray, freqs: np.ndarray, times: np.ndarray) -> str:
    try:
        peak_freqs = freqs[np.argmax(S_db, axis=0)]
        freq_range = peak_freqs.max() - peak_freqs.min()
        if freq_range < 5:
            return "CF (Constant Frequency)"
        elif freq_range > 30:
            start_freq = np.median(peak_freqs[: len(peak_freqs) // 4])
            end_freq   = np.median(peak_freqs[-len(peak_freqs) // 4 :])
            if start_freq > end_freq + 10:
                return "FM-steep (Steep Frequency Modulated)"
            elif end_freq > start_freq + 10:
                return "FM-ascending"
            else:
                return "FM-shallow (Shallow Frequency Modulated)"
        else:
            return "QCF (Quasi-Constant Frequency)"
    except Exception as e:
        print(f"Shape analysis error: {e}")
        return "Unknown"


def detect_pulse_duration(y: np.ndarray, sr: int, threshold_db: float = -40) -> tuple:
    try:
        envelope    = np.abs(signal.hilbert(y))
        envelope_db = 20 * np.log10(envelope + 1e-10)
        envelope_db = envelope_db - np.max(envelope_db)

        above_threshold = envelope_db > threshold_db
        pulse_starts    = np.where(np.diff(above_threshold.astype(int)) ==  1)[0]
        pulse_ends      = np.where(np.diff(above_threshold.astype(int)) == -1)[0]

        if len(pulse_starts) == 0 or len(pulse_ends) == 0:
            total_length = len(y) / sr * 1000
            return total_length / 2, total_length

        pulse_durations = []
        for start, end in zip(pulse_starts, pulse_ends[: len(pulse_starts)]):
            duration_ms = (end - start) / sr * 1000
            if duration_ms > 0.5:
                pulse_durations.append(duration_ms)

        if not pulse_durations:
            pulse_durations = [5.0]

        total_length      = (pulse_ends[-1] - pulse_starts[0]) / sr * 1000 if len(pulse_starts) > 0 else np.median(pulse_durations)
        avg_pulse_duration = np.median(pulse_durations)
        return avg_pulse_duration, total_length
    except Exception as e:
        print(f"Pulse detection error: {e}")
        return 5.0, 10.0


def extract_call_parameters(audio_path: Path) -> CallParameters:
    try:
        enhanced_params = extract_enhanced_call_parameters(audio_path)
        return CallParameters(
            start_frequency=enhanced_params.start_frequency,
            end_frequency=enhanced_params.end_frequency,
            peak_frequency=enhanced_params.peak_frequency,
            bandwidth=enhanced_params.bandwidth,
            intensity=enhanced_params.intensity,
            pulse_duration=enhanced_params.call_length,
            total_length=enhanced_params.call_length * enhanced_params.pulse_count,
            shape=enhanced_params.sonotype,
        )
    except Exception as e:
        print(f"Call parameter extraction error: {e}")
        traceback.print_exc()
        return CallParameters(
            start_frequency=0.0, end_frequency=0.0, peak_frequency=0.0,
            bandwidth=0.0, intensity=0.0, pulse_duration=0.0,
            total_length=0.0, shape="Unknown",
        )


def generate_spectrogram_for_model(audio_path: Path, output_path: Path) -> bool:
    """RAW spectrogram – no axes, no labels – fed directly to the CNN."""
    try:
        y, sr = librosa.load(audio_path, sr=None)
        if not np.any(y):
            raise ValueError("Empty audio")

        n_fft, hop_length = 2048, 512
        D     = librosa.stft(y, n_fft=n_fft, hop_length=hop_length)
        S_db  = librosa.amplitude_to_db(np.abs(D), ref=np.max)
        freqs = librosa.fft_frequencies(sr=sr, n_fft=n_fft) / 1000.0

        freq_mask = (freqs >= 10) & (freqs <= 250)
        if np.any(freq_mask):
            S_db = S_db[freq_mask, :]

        fig, ax = plt.subplots(figsize=(10, 8), facecolor="black")
        ax.set_facecolor("black")
        ax.imshow(S_db, aspect="auto", origin="lower", cmap="viridis")
        ax.set_axis_off()
        plt.subplots_adjust(left=0, right=1, bottom=0, top=1)
        plt.margins(0, 0)
        ax.xaxis.set_major_locator(plt.NullLocator())
        ax.yaxis.set_major_locator(plt.NullLocator())
        plt.savefig(output_path, bbox_inches="tight", pad_inches=0, dpi=150, facecolor="black")
        plt.close(fig)
        return True
    except Exception as e:
        print(f"Model spectrogram generation error: {e}")
        traceback.print_exc()
        return False


def generate_spectrogram_for_display(
    audio_path: Path,
    output_path: Path,
    theme_name: str = "dark_viridis",
    call_params: Optional[CallParameters] = None,
    species_list: Optional[List[SpeciesDetection]] = None,
) -> bool:
    """Labelled, themed spectrogram for the UI / PDF report."""
    try:
        theme = SPECTROGRAM_THEMES.get(theme_name, SPECTROGRAM_THEMES["dark_viridis"])
        y, sr = librosa.load(audio_path, sr=None)
        if not np.any(y):
            raise ValueError("Empty audio")

        n_fft, hop_length = 2048, 512
        D     = librosa.stft(y, n_fft=n_fft, hop_length=hop_length)
        S_db  = librosa.amplitude_to_db(np.abs(D), ref=np.max)
        freqs = librosa.fft_frequencies(sr=sr, n_fft=n_fft) / 1000.0
        # times = librosa.frames_to_time(np.arange(S_db.shape[1]), sr=sr, hop_length=hop_length) # Unused but fine to keep

        freq_mask = (freqs >= 10) & (freqs <= 250)
        if np.any(freq_mask):
            S_db  = S_db[freq_mask, :]
            freqs = freqs[freq_mask]

        fig, ax = plt.subplots(figsize=(12, 8), facecolor=theme["bg_color"])
        ax.set_facecolor(theme["bg_color"])

        # CHANGED: y_axis="linear" ensures frequency ticks are drawn based on y_coords
        img = librosa.display.specshow(
            S_db, sr=sr, hop_length=hop_length,
            x_axis="time", y_axis="linear", ax=ax, 
            y_coords=freqs, cmap=theme["cmap"],
        )

        # Title
        if species_list and len(species_list) > 0:
            top = species_list[0]
            title = f"{top.species} ({top.confidence:.1f}%)"
            if len(species_list) > 1:
                title += f" + {len(species_list)-1} more"
        else:
            title = "Bat Call Spectrogram"

        ax.set_title(title, color=theme["tick_color"], fontsize=16, fontweight="bold", pad=20)
        ax.set_xlabel("Time (s)",       color=theme["tick_color"], fontsize=13)
        ax.set_ylabel("Frequency (kHz)", color=theme["tick_color"], fontsize=13)
        ax.tick_params(colors=theme["tick_color"], labelsize=11)
        
        # Explicitly set limits to ensure the full range is visible
        ax.set_ylim([freqs.min(), freqs.max()])

        cbar = plt.colorbar(img, ax=ax, format="%+2.0f dB")
        cbar.ax.yaxis.set_tick_params(color=theme["tick_color"])
        plt.setp(plt.getp(cbar.ax.axes, "yticklabels"), color=theme["tick_color"])

        if call_params:
            param_text = (
                f"Peak: {call_params.peak_frequency} kHz | "
                f"BW: {call_params.bandwidth} kHz | "
                f"Duration: {call_params.pulse_duration} ms\n"
                f"Shape: {call_params.shape} | "
                f"Intensity: {call_params.intensity:.1f} dB"
            )
            ax.text(
                0.02, 0.98, param_text, transform=ax.transAxes,
                fontsize=10, verticalalignment="top",
                bbox=dict(boxstyle="round", facecolor=theme["bg_color"], alpha=0.85, edgecolor=theme["tick_color"]),
                color=theme["tick_color"],
            )

        plt.savefig(output_path, facecolor=fig.get_facecolor(), bbox_inches="tight", pad_inches=0.2, dpi=150)
        plt.close(fig)
        return True
    except Exception as e:
        print(f"Display spectrogram generation error: {e}")
        traceback.print_exc()
        return False

def slow_down_audio(input_path: Path, output_path: Path, factor: int = 10) -> bool:
    try:
        import soundfile as sf
        y, sr = librosa.load(input_path, sr=None)
        y_slow = librosa.effects.time_stretch(y, rate=1 / factor)
        sf.write(output_path, y_slow, sr)
        return True
    except Exception as e:
        print(f"Audio slowdown error: {e}")
        return False


# ──────────────────────────────────
# AI helpers (Groq / Llama)
# ──────────────────────────────────
def generate_ai_report(results: List[AnalysisResult], query: Optional[str] = None) -> str:
    if not groq_client:
        return "AI report generation is not available. Please configure GROQ_API_KEY."
    try:
        species_summary: Dict[str, list] = {}
        for result in results:
            for det in result.species_detected:
                species_summary.setdefault(det.species, []).append({
                    "confidence": det.confidence,
                    "filename":   result.original_filename,
                    "params":     result.call_parameters.dict(),
                })

        context = (
            f"You are an expert bat ecologist analyzing bat call data.\n\n"
            f"Total recordings analyzed: {len(results)}\n"
            f"Species detected: {list(species_summary.keys())}\n\n"
            f"Detailed detections:\n{json.dumps(species_summary, indent=2)}\n\n"
        )
        context += (
            f"User query: {query}"
            if query
            else "Provide a comprehensive ecological analysis including species diversity, call characteristics, and recommendations."
        )

        response = groq_client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system", "content": "You are an expert bat ecologist and bioacoustician."},
                {"role": "user",   "content": context},
            ],
            temperature=0.7,
            max_tokens=2000,
        )
        return response.choices[0].message.content
    except Exception as e:
        print(f"AI report generation error: {e}")
        return f"Error generating AI report: {str(e)}"


# ──────────────────────────────────
# PDF report builder
# ──────────────────────────────────
def generate_pdf_report(result: AnalysisResult, spectrogram_path: Path) -> io.BytesIO:
    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=letter, title=f"Bat Call Analysis – {result.original_filename}")

    elements = []
    styles   = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle", parent=styles["Heading1"], fontSize=24,
        textColor=HexColor("#2E86AB"), spaceAfter=30, alignment=TA_CENTER,
    )

    elements.append(Paragraph("Bat Call Analysis Report", title_style))
    elements.append(Paragraph(f"<b>File:</b> {result.original_filename}", styles["Normal"]))
    elements.append(Paragraph(
        f"<b>Analysis Date:</b> {datetime.fromtimestamp(result.timestamp).strftime('%Y-%m-%d %H:%M:%S')}",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 0.3 * inch))

    # Species table
    elements.append(Paragraph("<b>Species Detected:</b>", styles["Heading2"]))
    species_data = [["Rank", "Species", "Confidence"]]
    for det in result.species_detected[:10]:
        species_data.append([str(det.rank), det.species, f"{det.confidence:.2f}%"])

    species_table = Table(species_data, colWidths=[0.8 * inch, 4 * inch, 1.5 * inch])
    species_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#2E86AB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#FFFFFF")),
        ("ALIGN",     (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME",  (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",  (0, 0), (-1, 0),  12),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("GRID",     (0, 0), (-1, -1), 1, HexColor("#CCCCCC")),
    ]))
    elements.append(species_table)
    elements.append(Spacer(1, 0.3 * inch))

    # Parameters table
    elements.append(Paragraph("<b>Call Parameters:</b>", styles["Heading2"]))
    p = result.call_parameters
    param_data = [
        ["Parameter",       "Value"],
        ["Start Frequency", f"{p.start_frequency} kHz"],
        ["End Frequency",   f"{p.end_frequency} kHz"],
        ["Peak Frequency",  f"{p.peak_frequency} kHz"],
        ["Bandwidth",       f"{p.bandwidth} kHz"],
        ["Intensity",       f"{p.intensity:.1f} dB"],
        ["Pulse Duration",  f"{p.pulse_duration} ms"],
        ["Total Length",    f"{p.total_length} ms"],
        ["Shape",           p.shape],
    ]
    param_table = Table(param_data, colWidths=[2.5 * inch, 3.5 * inch])
    param_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#2E86AB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), HexColor("#FFFFFF")),
        ("ALIGN",    (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (0, -1),  "Helvetica-Bold"),
        ("GRID",     (0, 0), (-1, -1), 1, HexColor("#CCCCCC")),
    ]))
    elements.append(param_table)
    elements.append(Spacer(1, 0.3 * inch))

    # Spectrogram image
    if spectrogram_path.exists():
        elements.append(Paragraph("<b>Spectrogram:</b>", styles["Heading2"]))
        elements.append(RLImage(str(spectrogram_path), width=6 * inch, height=4 * inch, kind="proportional"))

    # In generate_pdf_report function...
    # ... after elements.append(Paragraph("Bat Call Analysis Report", title_style)) ...

    # Add Bat Image if available
    if result.species_detected:
        img_path = get_species_image_path(result.species_detected[0].species)
        if img_path:
            # Draw species image
            elements.append(Paragraph(f"<b>Detected: {result.species_detected[0].species}</b>", styles["Heading2"]))
            elements.append(RLImage(str(img_path), width=2*inch, height=2*inch, kind="proportional"))
            elements.append(Spacer(1, 0.2 * inch))
    
    # ... continue with the rest of your existing PDF logic ...

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ============================================
# API ENDPOINTS
# ============================================
@app.get("/")
async def root():
    return {
        "message": "🦇 Bat Call Analyzer API v3.0 – Supabase Backend",
        "features": [
            "Multi-species detection (28 species)",
            "Supabase Storage (bucket: bcit-bat)",
            "Supabase DB (analysis_results table)",
            "True background sync (zero blocking)",
            "Local-first architecture",
            "Batch processing",
            "AI-powered reports (Groq / Llama)",
        ],
        "status":           "online",
        "supabase_enabled": SUPABASE_ENABLED,
        "storage":          "local-first, supabase sync in background",
    }


@app.get("/health")
async def health_check():
    return {
        "status":            "healthy",
        "supabase_storage":  SUPABASE_ENABLED,
        "supabase_db":       SUPABASE_ENABLED,
        "groq_ai":           groq_client is not None,
        "model_loaded":      predictor.model is not None,
        "species_count":     len(predictor.classes),
    }


# ──────────────────────────────────
# Single-file audio analysis
# ──────────────────────────────────
@app.post("/api/analyze/audio", response_model=AnalysisResult)
async def analyze_audio(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    theme: str = Form("dark_viridis"),
    threshold: float = Form(0.01),
    max_freq: int = Form(250),
    batch_id: Optional[str] = Form(None),
):
    if not file.filename.lower().endswith(".wav"):
        raise HTTPException(400, "Only WAV files supported")

    file_id = str(uuid.uuid4())
    audio_path = UPLOADS_DIR / f"{file_id}.wav"

    try:
        print(f"\n🎵 Analyzing Single: {file.filename}")
        
        # SAVE IMMEDIATELY while file is open
        with open(audio_path, "wb") as f:
            shutil.copyfileobj(file.file, f)

        # Delegate to core logic
        return await _process_audio_internal(
            file_id, file.filename, audio_path, background_tasks, theme, threshold, batch_id
        )

    except Exception as e:
        if audio_path.exists(): audio_path.unlink()
        raise HTTPException(500, f"Analysis failed: {str(e)}")


# ──────────────────────────────────
# Single spectrogram-image analysis
# ──────────────────────────────────
@app.post("/api/analyze/spectrogram", response_model=AnalysisResult)
async def analyze_spectrogram(
    background_tasks: BackgroundTasks,
    file:      UploadFile = File(...),
    threshold: float      = Form(0.01),
):
    if not file.filename.lower().endswith((".png", ".jpg", ".jpeg")):
        raise HTTPException(400, "Only PNG/JPG supported")

    file_id   = str(uuid.uuid4())
    spec_path = SPECTROGRAMS_DIR / f"{file_id}.png"

    try:
        print(f"\n🖼️  Analyzing spectrogram: {file.filename}")
        with open(spec_path, "wb") as f:
            shutil.copyfileobj(file.file, f)

        species_list      = predictor.predict_multi_species(str(spec_path), threshold)
        species_image_url = get_species_image_url(species_list[0]["species"]) if species_list else None

        result = AnalysisResult(
            file_id=file_id,
            original_filename=file.filename,
            timestamp=int(time.time()),
            duration=0.0,
            sample_rate=0,
            species_detected=[SpeciesDetection(**s) for s in species_list],
            call_parameters=CallParameters(
                start_frequency=0, end_frequency=0, peak_frequency=0,
                bandwidth=0, intensity=0, pulse_duration=0,
                total_length=0, shape="N/A (Image Input)",
            ),
            spectrogram_url=f"/api/spectrograms/{file_id}",
            audio_url="",
            supabase_urls=None,
            processing_mode="Direct Image",
            display_theme="N/A",
            species_image_url=species_image_url,
            sync_status="pending",
        )

        result_dict = result.dict()
        with open(RESULTS_DIR / f"{file_id}.json", "w") as f:
            json.dump(result_dict, f, indent=2)

        if SUPABASE_ENABLED:
            background_tasks.add_task(
                async_supabase_upload_task,
                file_id, Path(), Path(), spec_path, result_dict,
            )

        print("✅ Analysis complete!\n")
        return result

    except Exception as e:
        print(f"❌ Analysis failed: {e}")
        if spec_path.exists():
            spec_path.unlink()
        raise HTTPException(500, f"Analysis failed: {str(e)}")
async def _process_audio_internal(
    file_id: str,
    original_filename: str,
    audio_path: Path,
    background_tasks: BackgroundTasks,
    theme: str,
    threshold: float,
    batch_id: Optional[str]
) -> AnalysisResult:
    """
    Core analysis logic that works on a LOCAL FILE PATH.
    Does NOT handle UploadFile objects.
    """
    spec_display_path = SPECTROGRAMS_DIR / f"{file_id}.png"
    spec_model_path = SPECTROGRAMS_DIR / f"{file_id}_model.png"
    slow_audio_path = UPLOADS_DIR / f"{file_id}_slow.wav"

    try:
        # 1. Extract Parameters
        call_params = extract_call_parameters(audio_path)
        
        # 2. Generate Model Spectrogram & Predict
        generate_spectrogram_for_model(audio_path, spec_model_path)
        species_list = predictor.predict_multi_species(str(spec_model_path), threshold)
        
        # 3. Generate Display Spectrogram
        generate_spectrogram_for_display(
            audio_path, spec_display_path, theme_name=theme,
            call_params=call_params,
            species_list=[SpeciesDetection(**s) for s in species_list],
        )

        # 4. Audio Processing
        slow_down_audio(audio_path, slow_audio_path, factor=10)
        y, sr = librosa.load(audio_path, sr=None)
        duration = librosa.get_duration(y=y, sr=sr)

        species_image_url = get_species_image_url(species_list[0]["species"]) if species_list else None

        # 5. Build Result
        result = AnalysisResult(
            file_id=file_id,
            original_filename=original_filename,
            timestamp=int(time.time()),
            duration=round(duration, 2),
            sample_rate=int(sr),
            species_detected=[SpeciesDetection(**s) for s in species_list],
            call_parameters=call_params,
            spectrogram_url=f"/api/spectrograms/{file_id}",
            audio_url=f"/api/audio/{file_id}/slow",
            supabase_urls=None,
            processing_mode="Bandpass 10-250kHz",
            display_theme=theme,
            species_image_url=species_image_url,
            sync_status="pending",
            batch_id=batch_id,
        )

        # 6. Save Local JSON
        result_dict = result.dict()
        with open(RESULTS_DIR / f"{file_id}.json", "w") as f:
            json.dump(result_dict, f, indent=2)

        # 7. Queue Supabase Upload
        if SUPABASE_ENABLED:
            background_tasks.add_task(
                async_supabase_upload_task,
                file_id, audio_path, slow_audio_path, spec_display_path, result_dict,
            )

        return result

    except Exception as e:
        print(f"❌ Internal Processing Failed: {e}")
        traceback.print_exc()
        raise e


def save_batch_to_supabase(batch_dict: dict) -> bool:
    """
    Save batch metadata to Supabase batch_analyses table.
    """
    if not SUPABASE_ENABLED:
        return False
    try:
        # Create a clean copy to avoid modifying the original dict
        row = batch_dict.copy()
        
        # Add sync timestamp
        row["synced_at"] = datetime.utcnow().isoformat()
        
        # Serialize list fields (like file_ids) to JSON strings if they aren't already
        if "file_ids" in row and isinstance(row["file_ids"], list):
            row["file_ids"] = json.dumps(row["file_ids"])
            
        # Ensure we don't send extra fields that aren't in the DB schema
        # (Optional: filter keys if your DB is strict, but upsert usually handles extras by ignoring or erroring depending on setup)
        
        supabase.table("batch_analyses").upsert(row).execute()
        print(f"  ✓ Batch metadata saved: {row['batch_id']}")
        return True
    except Exception as e:
        print(f"  ✗ Batch save failed: {e}")
        # Don't print full traceback for batch save to keep logs clean, just the error
        return False


@app.get("/api/batches")
async def get_all_batches():
    """
    Get a list of all batch summaries.
    Reads from the local 'batches' directory.
    """
    batches = []
    
    # Ensure directory exists
    if not BATCH_DIR.exists():
        BATCH_DIR.mkdir(parents=True, exist_ok=True)
        
    # Read all JSON files in the batches directory
    for batch_file in BATCH_DIR.glob("*.json"):
        try:
            with open(batch_file) as f:
                data = json.load(f)
                # Only add valid batch files
                if "batch_id" in data:
                    batches.append(data)
        except Exception as e:
            print(f"⚠ Failed to load batch metadata {batch_file}: {e}")
            continue
    
    # Sort by date (newest first)
    batches.sort(key=lambda x: x.get("created_at", 0), reverse=True)
    
    return {"batches": batches, "count": len(batches)}
def get_batch_from_supabase(batch_id: str) -> Optional[Dict]:
    """Fetch batch metadata from Supabase."""
    if not SUPABASE_ENABLED:
        return None
    try:
        resp = supabase.table("batch_analyses").select("*").eq("batch_id", batch_id).execute()
        if resp.data:
            row = resp.data[0]
            # Deserialize file_ids if it's stored as a string
            if "file_ids" in row and isinstance(row["file_ids"], str):
                try:
                    row["file_ids"] = json.loads(row["file_ids"])
                except json.JSONDecodeError:
                    pass # Keep as is if not valid JSON
            return row
        return None
    except Exception as e:
        print(f"❌ Supabase batch fetch failed for {batch_id}: {e}")
        return None

@app.get("/api/batches/{batch_id}")
async def get_batch_details(batch_id: str):
    """
    Get full details for a specific batch, including all its results.
    """
    batch_path = BATCH_DIR / f"{batch_id}.json"
    
    if not batch_path.exists():
        # Fallback: Try to query Supabase if local file is missing
        if SUPABASE_ENABLED:
            data = get_batch_from_supabase(batch_id)
            if data:
                return {"results": [], **data} # You might need logic to fetch results from Supabase too
        raise HTTPException(404, f"Batch {batch_id} not found")
    
    try:
        with open(batch_path) as f:
            batch_meta = json.load(f)
            
        # Collect the actual analysis results for the files in this batch
        results = []
        file_ids = batch_meta.get("file_ids", [])
        
        for fid in file_ids:
            res_path = RESULTS_DIR / f"{fid}.json"
            if res_path.exists():
                with open(res_path) as rf:
                    results.append(json.load(rf))
                    
        # Return the list of results + metadata
        return {
            "batch_id": batch_meta.get("batch_id"),
            "total_files": batch_meta.get("total_files"),
            "completed": batch_meta.get("completed"),
            "failed": batch_meta.get("failed"),
            "results": results  # <--- The frontend needs this array
        }
        
    except Exception as e:
        print(f"Error reading batch {batch_id}: {e}")
        raise HTTPException(500, f"Failed to load batch: {str(e)}")
# ──────────────────────────────────
# Batch analysis
# ──────────────────────────────────
@app.post("/api/analyze/batch/stream")
async def analyze_batch_stream(
    background_tasks: BackgroundTasks,
    files: List[UploadFile] = File(...),
    input_type: str = Form("audio"),
    theme: str = Form("dark_viridis"),
    threshold: float = Form(0.01),
    max_freq: int = Form(250),
):
    """
    Streams results back to the client as newline-delimited JSON.
    Saves files locally FIRST to prevent 'closed file' errors during streaming.
    """
    batch_id = str(uuid.uuid4())
    
    # ---------------------------------------------------------
    # STEP 1: Synchronously save all files to disk first.
    # This prevents the "I/O operation on closed file" error.
    # ---------------------------------------------------------
    saved_files_queue = []
    
    print(f"📥 Batch Upload: Receiving {len(files)} files...")
    
    try:
        for file in files:
            file_id = str(uuid.uuid4())
            
            # Determine correct extension and path
            if input_type == "audio":
                ext = ".wav" if file.filename.lower().endswith(".wav") else Path(file.filename).suffix
                save_path = UPLOADS_DIR / f"{file_id}{ext}"
            else:
                ext = Path(file.filename).suffix
                save_path = SPECTROGRAMS_DIR / f"{file_id}{ext}"
            
            # Write to disk immediately 
            with open(save_path, "wb") as f:
                shutil.copyfileobj(file.file, f)
            
            saved_files_queue.append({
                "file_id": file_id,
                "path": save_path,
                "original_filename": file.filename
            })
            
    except Exception as e:
        print(f"❌ Batch Upload Failed: {e}")
        # Clean up any files saved so far if upload fails
        for item in saved_files_queue:
            if item["path"].exists():
                item["path"].unlink()
        raise HTTPException(500, f"Failed to save batch upload: {str(e)}")

    # ---------------------------------------------------------
    # STEP 2: Generator processes the SAVED files on disk
    # ---------------------------------------------------------
    async def process_generator():
        # Send Start Signal
        batch_meta = {
            "batch_id": batch_id,
            "total_files": len(saved_files_queue),
            "created_at": int(time.time()),
            "status": "processing",
            "file_ids": [],
            "theme": theme,
            "input_type": input_type
        }
        
        yield json.dumps({"type": "batch_start", **batch_meta}) + "\n"
        
        completed_count = 0
        failed_count = 0
        
        for idx, item in enumerate(saved_files_queue):
            try:
                # Use the internal helper function we created previously
                # (Ensure _process_audio_internal is defined in your app.py)
                if input_type == "audio":
                    result = await _process_audio_internal(
                        file_id=item['file_id'],
                        original_filename=item['original_filename'],
                        audio_path=item['path'],
                        background_tasks=background_tasks,
                        theme=theme,
                        threshold=threshold,
                        batch_id=batch_id
                    )
                else:
                    # Logic for spectrograms (if applicable)
                    # result = await _process_spectrogram_internal(...)
                    continue 

                batch_meta["file_ids"].append(result.file_id)
                completed_count += 1
                
                # Stream Success Result
                yield json.dumps({
                    "type": "result", 
                    "index": idx + 1,
                    "total": len(saved_files_queue),
                    "data": result.dict()
                }) + "\n"
                
            except Exception as e:
                print(f"❌ Batch processing error for {item['original_filename']}: {e}")
                failed_count += 1
                
                # Stream Error
                yield json.dumps({
                    "type": "error", 
                    "index": idx + 1,
                    "filename": item['original_filename'], 
                    "error": str(e)
                }) + "\n"

        # Finalize Batch Metadata
        batch_meta["status"] = "completed"
        batch_meta["completed_count"] = completed_count
        batch_meta["failed_count"] = failed_count
        
        # Save Batch History to Disk
        try:
            with open(BATCH_DIR / f"{batch_id}.json", "w") as f:
                json.dump(batch_meta, f, indent=2)
        except Exception as e:
            print(f"Failed to save batch JSON: {e}")
            
        # Optional: Save Batch Metadata to Supabase
        if SUPABASE_ENABLED:
             try:
                 # Ensure 'batch_analyses' table exists in Supabase
                 save_batch_to_supabase(batch_meta) 
             except Exception as e:
                 print(f"Supabase batch save failed: {e}")

        yield json.dumps({"type": "batch_complete", "batch_id": batch_id}) + "\n"

    return StreamingResponse(process_generator(), media_type="application/x-ndjson")

@app.get("/api/download/batch/{batch_id}/excel")
async def download_batch_excel(batch_id: str):
    # 1. Load batch metadata
    batch_file = BATCH_DIR / f"{batch_id}.json"
    if not batch_file.exists():
        # Fallback: Try to find all files with this batch_id from Results dir
        results = []
        for f in RESULTS_DIR.glob("*.json"):
            data = json.load(open(f))
            if data.get("batch_id") == batch_id:
                results.append(AnalysisResult(**data))
    else:
        meta = json.load(open(batch_file))
        results = []
        for fid in meta["file_ids"]:
            if (RESULTS_DIR / f"{fid}.json").exists():
                data = json.load(open(RESULTS_DIR / f"{fid}.json"))
                results.append(AnalysisResult(**data))

    if not results:
        raise HTTPException(404, "No results found for this batch")

    excel_file = generate_excel_report(results, batch_id)
    
    filename = f"Bat_Analysis_Batch_{batch_id[:8]}.xlsx"
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/download/batch/{batch_id}/pdf")
async def download_batch_pdf(batch_id: str):
    # Similar logic to Excel, but generate a multi-page PDF
    # For brevity, this uses the existing single PDF generator
    # You would ideally create a `generate_merged_pdf` function
    pass

# ──────────────────────────────────
# Result retrieval / deletion
# ──────────────────────────────────
@app.get("/api/results")
async def get_all_results():
    """Local-first: reads JSON files from disk (instant)."""
    results = []
    for result_file in RESULTS_DIR.glob("*.json"):
        try:
            with open(result_file) as f:
                data = json.load(f)
                data.setdefault("file_id", result_file.stem)
                results.append(data)
        except Exception as e:
            print(f"❌ Failed to load {result_file}: {e}")
    results.sort(key=lambda x: x.get("timestamp", 0), reverse=True)
    return {"results": results, "source": "local", "count": len(results)}


@app.get("/api/results/{file_id}")
async def get_result(file_id: str):
    result_path = RESULTS_DIR / f"{file_id}.json"
    if result_path.exists():
        with open(result_path) as f:
            data = json.load(f)
            data.setdefault("file_id", file_id)
            return data
    # fallback to Supabase
    if SUPABASE_ENABLED:
        data = get_result_from_supabase(file_id)
        if data:
            return data
    raise HTTPException(404, f"Result {file_id} not found")


@app.delete("/api/results/{file_id}")
async def delete_result(file_id: str):
    print(f"\n🗑️  Deleting result: {file_id}")
    deleted_items: List[str] = []

    # ── Supabase DB ──
    if SUPABASE_ENABLED:
        if delete_result_from_supabase(file_id):
            deleted_items.append("supabase_row")
        # ── Supabase Storage ──
        for obj_path in [f"audio/{file_id}.wav", f"audio/{file_id}_slow.wav", f"spectrograms/{file_id}.png"]:
            if delete_from_supabase_storage(obj_path):
                deleted_items.append(f"storage_{obj_path}")

    # ── local files ──
    for ext in (".json", ".wav", ".png", "_slow.wav", "_model.png"):
        for directory in [RESULTS_DIR, UPLOADS_DIR, SPECTROGRAMS_DIR]:
            fp = directory / f"{file_id}{ext}"
            if fp.exists():
                fp.unlink()
                deleted_items.append(f"local_{fp.name}")

    print(f"✓ Deleted {len(deleted_items)} items\n")
    return {"message": "Result deleted successfully", "deleted_items": deleted_items, "count": len(deleted_items)}


# ──────────────────────────────────
# Static file serving
# ──────────────────────────────────
@app.get("/api/spectrograms/{file_id}")
async def get_spectrogram(file_id: str):
    p = SPECTROGRAMS_DIR / f"{file_id}.png"
    if not p.exists():
        raise HTTPException(404, "Spectrogram not found")
    return FileResponse(p, media_type="image/png")


@app.get("/api/audio/{file_id}/{speed}")
async def get_audio(file_id: str, speed: str = "normal"):
    suffix = "_slow" if speed == "slow" else ""
    p = UPLOADS_DIR / f"{file_id}{suffix}.wav"
    if not p.exists():
        raise HTTPException(404, "Audio not found")
    return FileResponse(p, media_type="audio/wav")


@app.get("/api/static/bat_species/{image_name}")
async def get_species_image(image_name: str):
    p = STATIC_DIR / image_name
    if p.exists():
        return FileResponse(p)
    if not image_name.endswith((".jpg", ".jpeg", ".png")):
        p2 = STATIC_DIR / f"{image_name}.jpg"
        if p2.exists():
            return FileResponse(p2)
    placeholder = STATIC_DIR / "placeholder.jpg"
    if placeholder.exists():
        return FileResponse(placeholder)
    raise HTTPException(404, "Image not found")


# ──────────────────────────────────
# Chat (non-streaming + streaming)
# ──────────────────────────────────
class ChatRequest(BaseModel):
    message:    str
    history:    List[Dict[str, Any]]
    statistics: Optional[Dict[str, Any]] = None

class ChatResponse(BaseModel):
    response: str


@app.post("/api/chat", response_model=ChatResponse)
async def chat_with_groq(request: ChatRequest):
    try:
        context = (
            "You are an expert bat bioacoustics assistant. You have access to the following bat call analysis data:\n\n"
            f"STATISTICS:\n{json.dumps(request.statistics, indent=2) if request.statistics else 'No statistics available'}\n\n"
            "RECENT ANALYSIS HISTORY (last 20 recordings):\n"
        )
        for i, result in enumerate(request.history[:20]):
            sd = result.get("species_detected", [{}])
            cp = result.get("call_parameters", {})
            context += (
                f"\nRecording {i+1}:\n"
                f"- Filename: {result.get('original_filename', 'Unknown')}\n"
                f"- Top Species: {sd[0].get('species', 'Unknown') if sd else 'Unknown'} "
                f"({sd[0].get('confidence', 0):.1f}% confidence)\n"
                f"- Peak Frequency: {cp.get('peak_frequency', 'N/A')} kHz\n"
                f"- Duration: {result.get('duration', 'N/A')}s\n"
                f"- Pulse Duration: {cp.get('pulse_duration', 'N/A')} ms\n"
                f"- Shape: {cp.get('shape', 'N/A')}\n"
            )
        context += "\n\nPlease answer questions about this data accurately and helpfully."

        completion = groq_client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system", "content": context},
                {"role": "user",   "content": request.message},
            ],
            temperature=0.7, max_tokens=1024, top_p=1, stream=False,
        )
        return ChatResponse(response=completion.choices[0].message.content)
    except Exception as e:
        raise HTTPException(500, detail=f"Chat error: {str(e)}")


@app.post("/api/chat/stream")
async def chat_with_groq_stream(request: ChatRequest):
    try:
        context = (
            "You are an expert bat bioacoustics assistant with access to analysis data:\n\n"
            f"STATISTICS: {json.dumps(request.statistics, indent=2) if request.statistics else 'No statistics'}\n\n"
            "Answer questions about bat calls, species detection, and analysis patterns."
        )

        async def generate():
            stream = groq_client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[
                    {"role": "system", "content": context},
                    {"role": "user",   "content": request.message},
                ],
                temperature=0.7, max_tokens=1024, stream=True,
            )
            for chunk in stream:
                if chunk.choices[0].delta.content:
                    yield chunk.choices[0].delta.content

        return StreamingResponse(generate(), media_type="text/plain")
    except Exception as e:
        raise HTTPException(500, detail=f"Streaming error: {str(e)}")


# ──────────────────────────────────
# Sync endpoints (manual triggers)
# ──────────────────────────────────
from supabase_sync import SupabaseResultsSync   # noqa: E402

sync_manager: Optional[SupabaseResultsSync] = None
if SUPABASE_ENABLED:
    sync_manager = SupabaseResultsSync(
        supabase_client=supabase,
        bucket=SUPABASE_BUCKET,
        results_dir=RESULTS_DIR,
        uploads_dir=UPLOADS_DIR,
        spectrograms_dir=SPECTROGRAMS_DIR,
    )
    print("✓ Supabase Sync Manager initialised")


@app.post("/api/sync/upload")
async def sync_upload_all(force: bool = False):
    if not sync_manager:
        raise HTTPException(503, "Sync manager not available")
    stats = sync_manager.sync_all_to_supabase(force=force)
    return {"message": "Upload sync complete", "stats": stats, "timestamp": int(time.time())}


@app.post("/api/sync/download")
async def sync_download_all(force: bool = False):
    if not sync_manager:
        raise HTTPException(503, "Sync manager not available")
    stats = sync_manager.sync_all_from_supabase(force=force)
    return {"message": "Download sync complete", "stats": stats, "timestamp": int(time.time())}


@app.post("/api/sync/bidirectional")
async def sync_bidirectional(force: bool = False):
    if not sync_manager:
        raise HTTPException(503, "Sync manager not available")
    stats = sync_manager.bidirectional_sync(force=force)
    return {"message": "Bidirectional sync complete", "stats": stats, "timestamp": int(time.time())}


@app.post("/api/sync/result/{file_id}/upload")
async def sync_single_upload(file_id: str, force: bool = False):
    if not sync_manager:
        raise HTTPException(503, "Sync manager not available")
    result_path = RESULTS_DIR / f"{file_id}.json"
    if not result_path.exists():
        raise HTTPException(404, f"Result {file_id} not found locally")
    status = sync_manager.sync_result_to_supabase(file_id, force=force)
    return {"message": f"Uploaded {file_id}", "file_id": file_id, "sync_status": status, "timestamp": int(time.time())}


@app.post("/api/sync/result/{file_id}/download")
async def sync_single_download(file_id: str, force: bool = False):
    if not sync_manager:
        raise HTTPException(503, "Sync manager not available")
    success = sync_manager.sync_result_from_supabase(file_id, force=force)
    if not success:
        raise HTTPException(404, f"Failed to download {file_id}")
    return {"message": f"Downloaded {file_id}", "file_id": file_id, "success": success, "timestamp": int(time.time())}


@app.get("/api/sync/status")
async def get_sync_status():
    if not sync_manager:
        return {"supabase_enabled": False, "sync_available": False}
    return sync_manager.get_sync_status()


@app.post("/api/sync/daemon/start")
async def start_sync_daemon(interval: int = 300):
    if not sync_manager:
        raise HTTPException(503, "Sync manager not available")
    sync_manager.start_sync_daemon(interval=interval)
    return {"message": "Sync daemon started", "interval_seconds": interval, "status": sync_manager.get_sync_status()}


@app.post("/api/sync/daemon/stop")
async def stop_sync_daemon():
    if not sync_manager:
        raise HTTPException(503, "Sync manager not available")
    sync_manager.stop_sync_daemon()
    return {"message": "Sync daemon stopped", "status": sync_manager.get_sync_status()}


# ──────────────────────────────────
# Downloads (CSV / PDF)
# ──────────────────────────────────
@app.get("/api/download/csv")
async def download_csv():
    results = []
    for result_file in RESULTS_DIR.glob("*.json"):
        try:
            with open(result_file) as f:
                results.append(json.load(f))
        except Exception:
            continue
    if not results:
        raise HTTPException(404, "No results found")

    csv_buffer = io.StringIO()
    fieldnames = [
        "file_id", "original_filename", 
        "recording_date", "recording_time",
        "analysis_timestamp", "analysis_date", "analysis_time",
        "top_species", "top_confidence", "total_species_detected",
        "start_frequency_khz", "end_frequency_khz", "peak_frequency_khz",
        "bandwidth_khz", "intensity_db", "pulse_duration_ms",
        "total_length_ms", "call_shape", "duration_seconds",
        "sample_rate_hz", "processing_mode", "display_theme", "sync_status",
    ]
    writer = csv.DictWriter(csv_buffer, fieldnames=fieldnames)
    writer.writeheader()

    for r in results:
        sd = r.get("species_detected", [])
        cp = r.get("call_parameters", {})
        ts = r.get("timestamp", 0)
        analysis_dt = datetime.fromtimestamp(ts)
        
        # Extract recording date/time from filename
        filename = r.get("original_filename", "")
        recording_date, recording_time = extract_datetime_from_filename(filename)
        
        writer.writerow({
            "file_id":                r.get("file_id", ""),
            "original_filename":     filename,
            "recording_date":        recording_date or "N/A",
            "recording_time":        recording_time or "N/A",
            "analysis_timestamp":    ts,
            "analysis_date":         analysis_dt.strftime("%Y-%m-%d"),
            "analysis_time":         analysis_dt.strftime("%H:%M:%S"),
            "top_species":           sd[0]["species"] if sd else "N/A",
            "top_confidence":        f"{sd[0]['confidence']:.2f}" if sd else "0",
            "total_species_detected": len(sd),
            "start_frequency_khz":   cp.get("start_frequency", 0),
            "end_frequency_khz":     cp.get("end_frequency", 0),
            "peak_frequency_khz":    cp.get("peak_frequency", 0),
            "bandwidth_khz":         cp.get("bandwidth", 0),
            "intensity_db":          cp.get("intensity", 0),
            "pulse_duration_ms":     cp.get("pulse_duration", 0),
            "total_length_ms":       cp.get("total_length", 0),
            "call_shape":            cp.get("shape", "N/A"),
            "duration_seconds":      r.get("duration", 0),
            "sample_rate_hz":        r.get("sample_rate", 0),
            "processing_mode":       r.get("processing_mode", ""),
            "display_theme":         r.get("display_theme", ""),
            "sync_status":           r.get("sync_status", "unknown"),
        })

    csv_buffer.seek(0)
    return StreamingResponse(
        io.BytesIO(csv_buffer.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=bat_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"},
    )


@app.get("/api/download/pdf/{file_id}")
async def download_pdf(file_id: str):
    result_path = RESULTS_DIR / f"{file_id}.json"
    if not result_path.exists():
        raise HTTPException(404, "Result not found")
    with open(result_path) as f:
        result_data = json.load(f)
    result    = AnalysisResult(**result_data)
    spec_path = SPECTROGRAMS_DIR / f"{file_id}.png"
    pdf_buffer = generate_pdf_report(result, spec_path)
    filename  = result.original_filename.replace(".wav", "").replace(".png", "")
    return StreamingResponse(
        pdf_buffer, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=bat_report_{filename}_{datetime.now().strftime('%Y%m%d')}.pdf"},
    )


# ──────────────────────────────────
# AI report & explain
# ──────────────────────────────────
@app.post("/api/report/ai")
async def generate_ai_analysis(request: AIReportRequest):
    if not groq_client:
        raise HTTPException(503, "AI report generation unavailable – configure GROQ_API_KEY")
    results = []
    for fid in request.file_ids:
        p = RESULTS_DIR / f"{fid}.json"
        if p.exists():
            with open(p) as f:
                results.append(AnalysisResult(**json.load(f)))
    if not results:
        raise HTTPException(404, "No valid results found")
    report = generate_ai_report(results, request.query)
    return {"report": report, "analyzed_files": len(results), "file_ids": request.file_ids, "query": request.query, "timestamp": int(time.time())}


class CallParams(BaseModel):
    start_frequency: float
    end_frequency:   float
    peak_frequency:  float
    bandwidth:       float
    pulse_duration:  float
    shape:           str

class ExplanationRequest(BaseModel):
    filename:         str
    species_detected: List[SpeciesDetection]
    call_parameters:  CallParams


@app.post("/api/explain")
async def explain_analysis(request: ExplanationRequest):
    if not request.species_detected:
        return {"explanation": "No species detected to analyze."}
    dominant = request.species_detected[0]
    prompt = (
        f"You are a bioacoustics expert. Analyze this bat echolocation recording:\n"
        f"- Target Species: {dominant.species} (Confidence: {dominant.confidence:.2f})\n"
        f"- Call Shape: {request.call_parameters.shape}\n"
        f"- Peak Frequency: {request.call_parameters.peak_frequency} kHz\n"
        f"- Bandwidth: {request.call_parameters.bandwidth} kHz\n"
        f"- Duration: {request.call_parameters.pulse_duration} ms\n\n"
        f"Provide a concise, 3-bullet point explanation of what these specific parameters indicate about the bat's behavior "
        f"(e.g., search phase vs feeding buzz, environment type). Do not lecture on generalities; focus on these specific numbers."
    )
    try:
        chat_completion = groq_client.chat.completions.create(
            messages=[
                {"role": "system", "content": "You are a concise scientific assistant specializing in Chiroptera (bats) and acoustic signal processing."},
                {"role": "user",   "content": prompt},
            ],
            model="llama-3.3-70b-versatile",
            temperature=0.5, max_tokens=300,
        )
        return {"explanation": chat_completion.choices[0].message.content}
    except Exception as e:
        print(f"Groq API Error: {e}")
        raise HTTPException(500, detail="AI processing failed")


# ──────────────────────────────────
# Meta endpoints
# ──────────────────────────────────
@app.get("/api/species")
async def get_species_list():
    return {
        "species": [{"name": s, "image_url": get_species_image_url(s)} for s in predictor.classes],
        "total":   len(predictor.classes),
    }


@app.get("/api/themes")
async def get_themes():
    return {
        "themes":  list(SPECTROGRAM_THEMES.keys()),
        "default": "dark_viridis",
        "descriptions": {
            "dark_viridis":      "Dark background with viridis colormap",
            "bright_plasma":     "Bright background with plasma colormap",
            "classic_grayscale": "Classic black and white",
            "inferno":           "Dark background with inferno colormap",
            "magma":             "Dark background with magma colormap",
            "jet":               "Bright background with jet colormap",
        },
    }


@app.get("/api/stats")
async def get_statistics():
    results = []
    for f in RESULTS_DIR.glob("*.json"):
        try:
            with open(f) as fh:
                results.append(json.load(fh))
        except Exception:
            continue

    total          = len(results)
    species_counts: Dict[str, int] = {}
    total_duration = 0.0
    processing_modes: Dict[str, int] = {}
    themes_count:     Dict[str, int] = {}
    sync_statuses:    Dict[str, int] = {}

    for r in results:
        total_duration += r.get("duration", 0)
        for d in r.get("species_detected", []):
            species_counts[d["species"]] = species_counts.get(d["species"], 0) + 1
        processing_modes[r.get("processing_mode", "Unknown")] = processing_modes.get(r.get("processing_mode", "Unknown"), 0) + 1
        themes_count[r.get("display_theme", "Unknown")]        = themes_count.get(r.get("display_theme", "Unknown"), 0) + 1
        sync_statuses[r.get("sync_status", "unknown")]         = sync_statuses.get(r.get("sync_status", "unknown"), 0) + 1

    top_species = sorted(species_counts.items(), key=lambda x: x[1], reverse=True)[:10]

    return {
        "total_analyses":          total,
        "total_duration_seconds":  round(total_duration, 2),
        "total_duration_hours":    round(total_duration / 3600, 2),
        "unique_species_detected": len(species_counts),
        "top_species":             [{"species": s, "count": c, "percentage": round(c / total * 100, 1)} for s, c in top_species] if total else [],
        "average_duration_seconds": round(total_duration / total, 2) if total else 0,
        "processing_modes":        processing_modes,
        "popular_themes":          themes_count,
        "sync_statuses":           sync_statuses,
        "storage_type":            "local-first",
        "supabase_enabled":        SUPABASE_ENABLED,
    }


@app.get("/api/health/detailed")
async def detailed_health_check():
    local_results = len(list(RESULTS_DIR.glob("*.json")))
    sync_statuses = {"pending": 0, "syncing": 0, "synced": 0, "failed": 0}
    for rf in RESULTS_DIR.glob("*.json"):
        try:
            with open(rf) as f:
                status = json.load(f).get("sync_status", "unknown")
                if status in sync_statuses:
                    sync_statuses[status] += 1
        except Exception:
            pass

    return {
        "status":    "healthy",
        "timestamp": int(time.time()),
        "services": {
            "api":              "online",
            "supabase_storage": "enabled" if SUPABASE_ENABLED else "disabled",
            "supabase_db":      "connected" if SUPABASE_ENABLED else "disconnected",
            "groq_ai":          "enabled" if groq_client is not None else "disabled",
            "model":            "loaded" if predictor.model is not None else "not_loaded",
        },
        "data": {
            "local_results":  local_results,
            "sync_statuses":  sync_statuses,
            "total_species":  len(predictor.classes),
        },
        "storage_mode": "local-first with true background supabase sync",
    }


# ============================================
# LIFECYCLE HOOKS
# ============================================
@app.on_event("startup")
async def startup_event():
    print("\n" + "=" * 70)
    print("🦇  BAT CALL ANALYZER API v3.0 – SUPABASE BACKEND")
    print("=" * 70)
    print(f"✓ Model loaded:        {predictor.model is not None}")
    print(f"✓ Total species:       {len(predictor.classes)}")
    print(f"✓ Supabase Storage:    {'ENABLED' if SUPABASE_ENABLED else 'DISABLED'}  (bucket: {SUPABASE_BUCKET})")
    print(f"✓ Supabase DB:         {'ENABLED' if SUPABASE_ENABLED else 'DISABLED'}")
    print(f"✓ Sync Manager:        {'ENABLED' if sync_manager else 'DISABLED'}")
    print(f"✓ Groq AI:             {'ENABLED' if groq_client is not None else 'DISABLED'}")
    print("=" * 70)

    if sync_manager:
        print("\n🔄 Starting automatic sync daemon (5 min intervals)…")
        sync_manager.start_sync_daemon(interval=300)
        print("✓ Sync daemon started")

    print("\n🚀 API IS READY!")
    print(f"📚 Docs:     http://localhost:8000/docs")
    print(f"💾 Storage:  local-first (instant responses)")
    print(f"🔄 Sync:     Supabase uploads happen AFTER response sent")
    print("=" * 70 + "\n")


@app.on_event("shutdown")
async def shutdown_event():
    if sync_manager and sync_manager.daemon_running:
        print("\n🛑 Stopping sync daemon…")
        sync_manager.stop_sync_daemon()
        print("✅ Daemon stopped\n")


# ============================================
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="info")